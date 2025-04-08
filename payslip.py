import pandas as pd
from fpdf import FPDF
import os
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
EMAIL_ADDRESS = os.getenv("EMAIL_USER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASS")

# Step 1: Read employee data from Excel
try:
    import openpyxl  # Ensure the engine is available
    df = pd.read_excel("employees.xlsx", engine="openpyxl")
    if df.isnull().any().any():
        raise ValueError("Missing data detected in employees.xlsx")
except Exception as e:
    print(f"Error reading Excel file: {e}")
    exit()

# Step 2: Calculate Net Salary
df['Net Salary'] = df['Basic Salary'] + df['Allowances'] - df['Deductions']

# Step 3: Create payslips folder
os.makedirs("payslips", exist_ok=True)

# Step 4: Generate PDF payslips with professional layout
for index, row in df.iterrows():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", 'B', 16)
    pdf.cell(200, 10, txt="Payslip", ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Helvetica", size=12)
    pdf.cell(100, 10, txt=f"Employee ID: {row['Employee ID']}", ln=True)
    pdf.cell(100, 10, txt=f"Name: {row['Name']}", ln=True)
    pdf.ln(10)

    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(100, 10, txt="Salary Details:", ln=True)
    pdf.set_font("Helvetica", size=12)
    pdf.cell(100, 10, txt=f"Basic Salary: ${row['Basic Salary']:.2f}", ln=True)
    pdf.cell(100, 10, txt=f"Allowances: ${row['Allowances']:.2f}", ln=True)
    pdf.cell(100, 10, txt=f"Deductions: ${row['Deductions']:.2f}", ln=True)
    pdf.cell(100, 10, txt=f"Net Salary: ${row['Net Salary']:.2f}", ln=True)

    pdf.ln(20)
    pdf.set_font("Helvetica", style="I", size=10)
    pdf.cell(100, 10, txt="This is a system-generated payslip.", ln=True)

    file_path = f"payslips/{row['Employee ID']}.pdf"
    pdf.output(file_path)

# Step 5: Email the payslips
for index, row in df.iterrows():
    try:
        msg = EmailMessage()
        msg['Subject'] = "Your Payslip for This Month"
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = row['Email']
        msg.set_content(f"Dear {row['Name']},\n\nPlease find your payslip attached.\n\nRegards,\nHR Team")

        file_path = f"payslips/{row['Employee ID']}.pdf"
        with open(file_path, 'rb') as f:
            file_data = f.read()
            msg.add_attachment(file_data, maintype='application', subtype='pdf', filename=os.path.basename(file_path))

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)
        print(f"Payslip sent to {row['Email']}")
    except Exception as e:
        print(f"Error sending email to {row['Email']}: {e}")
